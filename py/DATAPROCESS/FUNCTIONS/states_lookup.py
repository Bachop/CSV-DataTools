# -*- coding: utf-8 -*-
"""
状态变量检测模块
用于检测状态变量为特定值时对应的传感器值变化情况
"""

import numpy as np
from PyQt5.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QPushButton,
                             QLabel, QSplitter, QDesktopWidget, QWidget,
                             QSizePolicy, QFrame, QApplication, QComboBox,
                             QDialogButtonBox, QAction, QToolButton,
                             QGroupBox, QStyle, QMenu, QTextEdit,
                             QLineEdit, QFormLayout, QSpinBox, QDialog)
from PyQt5.QtCore import Qt, QEvent, QTimer
from PyQt5.QtGui import QFont, QIcon, QKeyEvent, QStandardItemModel, QStandardItem

from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import matplotlib.text

# 添加openpyxl导入
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False
    Workbook = None


class NoArrowLineEdit(QLineEdit):
    """自定义LineEdit，防止方向键事件被拦截"""
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
    
    def keyPressEvent(self, event):
        """处理键盘事件，将方向键事件传递给父窗口"""
        if event.key() in (Qt.Key_Up, Qt.Key_Down, Qt.Key_Left, Qt.Key_Right):
            # 将方向键事件传递给父窗口
            if self.parent():
                self.parent().keyPressEvent(event)
            event.accept()
            return
        # 其他按键按默认处理
        super().keyPressEvent(event)


class SegmentDropdownButton(QToolButton):
    """支持键盘导航的状态段选择下拉按钮"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.segment_menu = None
        self.current_index = -1
        self.actions = []
        # 启用鼠标跟踪，确保能接收到鼠标事件
        self.setMouseTracking(True)
        # 安装事件过滤器以处理鼠标滚轮事件
        self.installEventFilter(self)
        
    def setMenu(self, menu):
        """设置菜单并保存引用"""
        super().setMenu(menu)
        self.segment_menu = menu
        self.actions = menu.actions()
        
    def keyPressEvent(self, event):
        """处理键盘事件，支持方向键导航"""
        if self.segment_menu and self.segment_menu.isVisible():
            if event.key() == Qt.Key_Up or event.key() == Qt.Key_Left:
                self.navigate_up()
                event.accept()
                return
            elif event.key() == Qt.Key_Down or event.key() == Qt.Key_Right:
                self.navigate_down()
                event.accept()
                return
            elif event.key() in (Qt.Key_Return, Qt.Key_Enter):
                self.activate_current_action()
                event.accept()
                return
            elif event.key() == Qt.Key_Escape:
                self.segment_menu.hide()
                event.accept()
                return
                
        # 对于其他按键，调用父类处理
        super().keyPressEvent(event)
        
    def eventFilter(self, obj, event):
        """事件过滤器，用于处理鼠标滚轮事件"""
        # 检查事件是否为鼠标滚轮事件，且菜单未显示
        if event.type() == QEvent.Wheel and self.segment_menu and not self.segment_menu.isVisible():
            # 获取滚轮滚动方向
            delta = event.angleDelta().y()
            if delta > 0:  # 向上滚动
                if hasattr(self.parent(), 'navigate_segments'):
                    self.parent().navigate_segments(-1)  # 导航到上一个状态段
            elif delta < 0:  # 向下滚动
                if hasattr(self.parent(), 'navigate_segments'):
                    self.parent().navigate_segments(1)   # 导航到下一个状态段
                    
            return True  # 事件已处理
            
        return super().eventFilter(obj, event)
        
    def navigate_up(self):
        """向上导航到上一个菜单项"""
        if not self.actions:
            return
            
        if self.current_index == -1:
            # 如果当前没有选中项，选中最后一个
            self.current_index = len(self.actions) - 1
        else:
            # 向上移动，如果在顶部则跳转到底部
            self.current_index = (self.current_index - 1) % len(self.actions)
            
        self.update_menu_selection()
        
    def navigate_down(self):
        """向下导航到下一个菜单项"""
        if not self.actions:
            return
            
        if self.current_index == -1:
            # 如果当前没有选中项，选中第一个
            self.current_index = 0
        else:
            # 向下移动，如果在底部则跳转到顶部
            self.current_index = (self.current_index + 1) % len(self.actions)
            
        self.update_menu_selection()
        
    def update_menu_selection(self):
        """更新菜单选择状态"""
        # 清除之前的选择状态
        for i, action in enumerate(self.actions):
            if i == self.current_index:
                # 设置为选中状态样式
                action.setFont(QFont(action.font().family(), action.font().pointSize(), QFont.Bold))
            else:
                # 恢复正常样式
                action.setFont(QFont(action.font().family(), action.font().pointSize(), QFont.Normal))
                
    def activate_current_action(self):
        """激活当前选中的菜单项"""
        if 0 <= self.current_index < len(self.actions):
            action = self.actions[self.current_index]
            action.trigger()
            if self.segment_menu:
                self.segment_menu.hide()
                
    def showMenu(self):
        """显示菜单时重置当前索引"""
        self.current_index = -1
        if self.segment_menu:
            # 重置所有菜单项样式
            for action in self.actions:
                action.setFont(QFont(action.font().family(), action.font().pointSize(), QFont.Normal))
        super().showMenu()


class StatesColumnSelectionDialog(QDialog):
    """状态变量检测列选择对话框"""
    
    def __init__(self, column_names, parent=None):
        super().__init__(parent)
        self.column_names = column_names
        self.state_column = None
        # sensor_column now may be a list of indices
        self.sensor_column = None
        
        self.setWindowTitle("状态变量检测 - 列选择")
        self.setModal(True)
        self.resize(600, 200)
        self.center()
        
        self.setup_ui()
    
    def center(self):
        """将窗口居中显示"""
        if self.parent():
            parent_geo = self.parent().frameGeometry()
            self.move(
                parent_geo.center().x() - self.width() // 2,
                parent_geo.center().y() - self.height() // 2
            )
        else:
            qr = self.frameGeometry()
            cp = QDesktopWidget().availableGeometry().center()
            qr.moveCenter(cp)
            self.move(qr.topLeft())
    
    def setup_ui(self):
        """创建UI元素"""
        layout = QVBoxLayout()
        
        # 状态变量列选择
        state_layout = QHBoxLayout()
        state_label = QLabel("状态变量列:")
        self.state_combo = QComboBox()

        for i, name in enumerate(self.column_names):
            self.state_combo.addItem(f"{name} (列{i+1})")  # 添加列索引备注

        state_layout.addWidget(state_label)
        state_layout.addWidget(self.state_combo)

        layout.addLayout(state_layout)
        
        # 传感器值列选择：先输入需要多少个传感器下拉框（n），然后显示 n 个下拉框
        sensor_layout = QHBoxLayout()
        sensor_count_label = QLabel("传感器列数:")
        self.sensor_count_spin = QSpinBox()
        self.sensor_count_spin.setRange(1, max(1, len(self.column_names)))
        self.sensor_count_spin.setValue(1)
        self.sensor_count_spin.valueChanged.connect(self._on_sensor_count_changed)

        sensor_layout.addWidget(sensor_count_label)
        sensor_layout.addWidget(self.sensor_count_spin)

        layout.addLayout(sensor_layout)

        # 放置动态下拉框的容器
        self.sensor_combos_container = QWidget()
        self.sensor_combos_layout = QHBoxLayout(self.sensor_combos_container)
        self.sensor_combos_layout.setContentsMargins(0, 0, 0, 0)
        self.sensor_combos_layout.setSpacing(4)
        self.sensor_combos = []

        # 初始化一个下拉框
        self._build_sensor_combos(1)

        layout.addWidget(self.sensor_combos_container)
        
        # 确定/取消按钮
        button_box = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel,
            Qt.Horizontal, self
        )
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
        
        self.setLayout(layout)

    def _on_sensor_count_changed(self, n):
        """响应传感器数目变化，重建下拉框"""
        try:
            n = int(n)
        except Exception:
            return
        self._build_sensor_combos(n)

    def _build_sensor_combos(self, n):
        """构建 n 个下拉框并加入到布局中"""
        # 清理已有
        for i in reversed(range(self.sensor_combos_layout.count())):
            w = self.sensor_combos_layout.itemAt(i).widget()
            if w:
                w.setParent(None)
        self.sensor_combos = []

        # 当前已选项（用于禁用）
        current_selected = []
        # 保留之前选择
        try:
            for c in getattr(self, 'sensor_combos', []):
                current_selected.append(c.currentIndex())
        except Exception:
            current_selected = []

        for i in range(n):
            combo = QComboBox()
            combo.setSizeAdjustPolicy(QComboBox.AdjustToContents)
            model = QStandardItemModel()
            # 填充项
            for idx, name in enumerate(self.column_names):
                display = f"{name} (列{idx+1})"
                item = QStandardItem(display)
                item.setData(idx)
                model.appendRow(item)
            combo.setModel(model)
            # 设定当前索引，如果之前有保存，则尝试恢复
            if i < len(current_selected) and isinstance(current_selected[i], int) and 0 <= current_selected[i] < len(self.column_names):
                combo.setCurrentIndex(current_selected[i])
            else:
                combo.setCurrentIndex(0)

            combo.currentIndexChanged.connect(self._on_sensor_combo_changed)
            self.sensor_combos_layout.addWidget(combo)
            self.sensor_combos.append(combo)

        # 同步禁用已被其他下拉框选中的项
        self._refresh_combo_models()

    def _on_sensor_combo_changed(self, _):
        """当任一下拉框值改变时，刷新其它下拉框的可用项（禁用已选项）"""
        self._refresh_combo_models()

    def _refresh_combo_models(self):
        """刷新所有下拉框的model，禁用已被其他下拉框选择的项"""
        selected = [c.currentIndex() for c in self.sensor_combos if c.currentIndex() >= 0]
        for i, combo in enumerate(self.sensor_combos):
            model = combo.model()
            # 遍历model中的项，设置enabled根据是否为其他combo选择
            for row in range(model.rowCount()):
                item = model.item(row)
                if item is None:
                    continue
                idx = row
                # 如果该项在其他combo中被选中，则禁用（但保留自身选中的项可选）
                other_selected = [s for j, s in enumerate(selected) if j != i]
                if idx in other_selected:
                    # 禁用
                    item.setEnabled(False)
                else:
                    item.setEnabled(True)
    
    def set_defaults(self, state_column=None, sensor_column=None):
        """
        设置默认选中的列
        
        Args:
            state_column (int or str): 状态变量列的索引或列名
            sensor_column (int or str): 传感器值列的索引或列名
        """
        if state_column is not None:
            if isinstance(state_column, int) and 0 <= state_column < self.state_combo.count():
                self.state_combo.setCurrentIndex(state_column)
            elif isinstance(state_column, str) and state_column in self.column_names:
                index = self.column_names.index(state_column)
                self.state_combo.setCurrentIndex(index)
        
        if sensor_column is not None:
            # sensor_column can be an int, str or a list of these
            indices = []
            if isinstance(sensor_column, (list, tuple)):
                for sc in sensor_column:
                    if isinstance(sc, int) and 0 <= sc < len(self.column_names):
                        indices.append(sc)
                    elif isinstance(sc, str) and sc in self.column_names:
                        indices.append(self.column_names.index(sc))
            else:
                if isinstance(sensor_column, int) and 0 <= sensor_column < len(self.column_names):
                    indices = [sensor_column]
                elif isinstance(sensor_column, str) and sensor_column in self.column_names:
                    indices = [self.column_names.index(sensor_column)]

            if indices:
                n = len(indices)
                self.sensor_count_spin.setValue(n)
                # 确保下拉框已创建
                self._build_sensor_combos(n)
                # 设置每个下拉框的选中项
                for i, idx in enumerate(indices):
                    if i < len(self.sensor_combos):
                        self.sensor_combos[i].setCurrentIndex(idx)
    
    def accept(self):
        """确认选择"""
        self.state_column = self.state_combo.currentIndex()
        # 从动态下拉框中获取选择
        sensor_indices = []
        for combo in getattr(self, 'sensor_combos', []):
            idx = combo.currentIndex()
            if idx is not None and idx >= 0:
                sensor_indices.append(idx)

        # 去重并保持顺序
        seen = set()
        uniq = []
        for x in sensor_indices:
            if x not in seen:
                seen.add(x)
                uniq.append(x)

        self.sensor_column = uniq

        # 检查是否选择了相同的列
        if self.state_column in self.sensor_column:
            from PyQt5.QtWidgets import QMessageBox
            QMessageBox.warning(self, "警告", "状态变量列和传感器值列不能相同，请重新选择。")
            return
        if not self.sensor_column:
            from PyQt5.QtWidgets import QMessageBox
            QMessageBox.warning(self, "警告", "请至少选择一列传感器值。")
            return

        super().accept()

class StatesLookupWindow(QDialog):
    """状态变量检测窗口"""
    
    def __init__(self, data, state_column, capacitor_columns, parent=None, original_file_path=None):
        super().__init__(parent)
        self.data = data
        self.state_column = state_column
        # capacitor_columns 可以是单个索引或列表
        if isinstance(capacitor_columns, (list, tuple)):
            self.capacitor_columns = list(capacitor_columns)
        else:
            self.capacitor_columns = [capacitor_columns]

        # 保存表头（用于图例显示）
        self.headers = self.data[0] if self.data and len(self.data) > 0 else []

        # 以状态列有值的行作为基准，保证各序列对齐
        rows_with_state = [row for row in self.data[1:] if row and row[self.state_column]]
        self.state_data = [float(row[self.state_column]) for row in rows_with_state]

        # 为每个传感器列构建对齐的数据序列，缺失值用 np.nan 填充
        self.capacitor_data_list = []
        for col in self.capacitor_columns:
            series = []
            for row in rows_with_state:
                try:
                    v = row[col]
                    series.append(float(v) if v != '' and v is not None else np.nan)
                except Exception:
                    series.append(np.nan)
            self.capacitor_data_list.append(series)
        
        # 分析状态数据，找出连续为1的段
        self.state_segments = self._find_state_segments()
        
        # 初始化当前段的x_range
        self.current_segment_x_range = []
        
        # 默认左右检测点数
        self.left_points = 30
        self.right_points = 30
        # 当前选中的传感器列（实际表格列索引）
        self.selected_curves = set()
        
        # 当前选中的状态段索引
        self.current_segment_index = -1
        
        # 保存原始文件路径，用于生成Excel文件名
        self.original_file_path = original_file_path
        
        self.setWindowTitle("状态变量检测")
        self.resize(1600, 900)  # 调整默认大小以适应不同分辨率
        self.setWindowFlags(Qt.Window | Qt.WindowMinMaxButtonsHint | Qt.WindowCloseButtonHint)  # 支持最大化最小化
        
        # 设置焦点策略，确保窗口能接收键盘事件
        self.setFocusPolicy(Qt.StrongFocus)
        self.setAttribute(Qt.WA_KeyCompression, False)
        
        # 初始化Matplotlib支持中文
        self.setup_matplotlib_chinese_support()
        
        # 居中显示窗口
        self.center()
        
        # 创建UI
        self.setup_ui()
        
        # 如果有状态段，显示第一个
        if self.state_segments:
            self.show_segment(0)
        
        # 在显示后强制刷新界面
        self.show()
        self.raise_()
        self.activateWindow()
    
    def center(self):
        """将窗口居中显示"""
        # 始终相对于屏幕居中
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())
    
    def setup_matplotlib_chinese_support(self):
        """设置Matplotlib支持中文"""
        import matplotlib as mpl
        # 检测操作系统
        system = QApplication.instance().platformName() if QApplication.instance() else ""
        
        # 设置中文字体
        if system == "windows":
            mpl.rcParams['font.sans-serif'] = ['SimHei']  # Windows黑体
        elif system == "darwin":  # macOS
            mpl.rcParams['font.sans-serif'] = ['Arial Unicode MS']
        else:  # Linux
            mpl.rcParams['font.sans-serif'] = ['WenQuanYi Micro Hei']
        
        # 解决负号显示问题
        mpl.rcParams['axes.unicode_minus'] = False
    
    def _find_state_segments(self):
        """
        查找状态数据中连续为1的段
        返回: [{"start": 起始索引, "end": 结束索引, "length": 长度}, ...]
        """
        segments = []
        if not self.state_data:
            return segments
            
        start = None
        for i, value in enumerate(self.state_data):
            # 检查是否是状态1的开始
            if value == 1 and start is None:
                start = i
            # 检查是否是状态1的结束
            elif value != 1 and start is not None:
                segments.append({
                    "start": start,
                    "end": i - 1,
                    "length": i - start
                })
                start = None
        
        # 处理最后一个段
        if start is not None:
            segments.append({
                "start": start,
                "end": len(self.state_data) - 1,
                "length": len(self.state_data) - start
            })
            
        return segments
    
    def setup_ui(self):
        """创建UI元素"""
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(5, 5, 5, 5)  # 减小边距以适应小分辨率
        main_layout.setSpacing(5)

        # 创建主水平分割器
        main_splitter = QSplitter(Qt.Horizontal)
        main_splitter.setChildrenCollapsible(False)  # 防止子控件被压缩到不可见

        # 左侧图表区域 - 使用垂直分割器
        left_widget = QSplitter(Qt.Vertical)
        left_widget.setChildrenCollapsible(False)  # 防止子控件被压缩到不可见

        # 第一个splitter - 显示所有数据曲线
        chart1_container = QWidget()
        chart1_layout = QVBoxLayout(chart1_container)
        chart1_layout.setContentsMargins(2, 2, 2, 2)
        chart1_layout.setSpacing(2)

        self.figure1 = Figure(figsize=(5, 3), dpi=100)  # 设置合适的默认大小
        self.canvas1 = FigureCanvas(self.figure1)
        self.canvas1.setMinimumSize(200, 150)  # 设置最小尺寸
        chart1_layout.addWidget(self.canvas1)

        # 第二个splitter - 显示当前选中段及其左右30个点
        chart2_container = QWidget()
        chart2_layout = QVBoxLayout(chart2_container)
        chart2_layout.setContentsMargins(2, 2, 2, 2)
        chart2_layout.setSpacing(2)

        self.figure2 = Figure(figsize=(5, 3), dpi=100)  # 设置合适的默认大小
        self.canvas2 = FigureCanvas(self.figure2)
        self.canvas2.setMinimumSize(200, 150)  # 设置最小尺寸
        chart2_layout.addWidget(self.canvas2)

        # 添加到左侧分割器
        left_widget.addWidget(chart1_container)
        left_widget.addWidget(chart2_container)
        left_widget.setSizes([400, 400])  # 设置合理的默认大小比例

        # 右侧控件区域
        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        right_layout.setContentsMargins(3, 3, 3, 3)
        right_layout.setSpacing(5)

        # 使用垂直分割器组织右侧控件，支持用户拖拽调节高度
        right_splitter = QSplitter(Qt.Vertical)
        right_splitter.setChildrenCollapsible(False)  # 防止子控件被压缩到不可见
        
        # 状态段选择下拉按钮
        segment_group = QGroupBox("状态段选择")
        segment_layout = QVBoxLayout(segment_group)

        # 创建下拉按钮
        self.segment_dropdown = SegmentDropdownButton()
        self.segment_dropdown.setText("选择状态段")
        self.segment_dropdown.setPopupMode(QToolButton.InstantPopup)
        self.segment_dropdown.setToolButtonStyle(Qt.ToolButtonTextOnly)

        # 创建下拉菜单
        self.segment_menu = QMenu(self.segment_dropdown)
        self.segment_dropdown.setMenu(self.segment_menu)

        # 创建状态段菜单项
        self.segment_actions = []
        for i, segment in enumerate(self.state_segments):
            action = QAction(f"状态段 {i+1}", self)
            action.triggered.connect(lambda checked, x=i: self.show_segment(x))
            self.segment_menu.addAction(action)
            self.segment_actions.append(action)

        segment_layout.addWidget(self.segment_dropdown)

        # 左右检测点数设置
        points_group = QGroupBox("检测点数设置")
        points_layout = QFormLayout(points_group)
        
        # 使用自定义LineEdit以防止方向键被拦截
        self.left_points_edit = NoArrowLineEdit(str(self.left_points))
        self.right_points_edit = NoArrowLineEdit(str(self.right_points))
        
        points_layout.addRow("左侧检测点数:", self.left_points_edit)
        points_layout.addRow("右侧检测点数:", self.right_points_edit)
        
        # 添加应用按钮
        apply_button = QPushButton("应用设置")
        apply_button.clicked.connect(self.apply_points_settings)
        points_layout.addRow(apply_button)
        
        # 添加保存Excel按钮
        if XLSX_AVAILABLE:
            save_excel_button = QPushButton("保存统计量到Excel")
            save_excel_button.clicked.connect(self.save_stats_to_excel)
            points_layout.addRow(save_excel_button)

        # 统计信息区域
        stats_group = QGroupBox("状态统计量")
        stats_layout = QVBoxLayout(stats_group)

        self.stats_label = QTextEdit()
        self.stats_label.setReadOnly(True)
        self.stats_label.setAlignment(Qt.AlignTop)
        self.stats_label.setHtml("请选择一个状态段以查看统计信息")
        self.stats_label.setMinimumSize(200, 400)  # 设置最小尺寸
        self.stats_label.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)
        stats_layout.addWidget(self.stats_label)

        # 曲线选择区域（用于多曲线时专注分析）
        curves_group = QGroupBox("曲线选择")
        curves_layout = QVBoxLayout(curves_group)
        curves_layout.setContentsMargins(4, 4, 4, 4)
        self.curve_buttons_widget = QWidget()
        self.curve_buttons_layout = QVBoxLayout(self.curve_buttons_widget)
        self.curve_buttons_layout.setContentsMargins(0, 0, 0, 0)
        curves_layout.addWidget(self.curve_buttons_widget)
        
        # 添加到右侧分割器
        right_splitter.addWidget(curves_group)
        right_splitter.addWidget(segment_group)
        right_splitter.addWidget(points_group)
        right_splitter.addWidget(stats_group)
        
        # 设置分割器的默认大小比例
        right_splitter.setSizes([30, 30, 30, 300])

        # 将分割器添加到右侧布局
        right_layout.addWidget(right_splitter)
        
        # 使用提示标签（鼠标中键平移，滚轮缩放，按 R 复位）
        try:
            from PyQt5.QtWidgets import QLabel
            self.hint_label = QLabel("提示: 左键点击拖拽平移，右键点击显示坐标，滚轮进行水平缩放，R 键恢复视图")
            self.hint_label.setWordWrap(True)
            self.hint_label.setStyleSheet('color: gray; font-size: 16px;')
            right_layout.addWidget(self.hint_label)
        except Exception:
            pass

        # 添加到主分割器
        main_splitter.addWidget(left_widget)
        main_splitter.addWidget(right_widget)
        main_splitter.setSizes([850, 350])  # 设置主分割器的默认大小比例

        main_layout.addWidget(main_splitter)

        # 创建图表子图
        self.ax1_cap = self.figure1.add_subplot(211)  # 电容值曲线
        self.ax1_state = self.figure1.add_subplot(212)  # 状态变量曲线

        self.ax2_cap = self.figure2.add_subplot(211)  # 选中段及周边电容值曲线
        self.ax2_state = self.figure2.add_subplot(212)  # 选中段及周边状态变量曲线

        # 连接鼠标事件
        self.canvas1.mpl_connect('button_press_event', self.on_mouse_press)
        self.canvas2.mpl_connect('button_press_event', self.on_mouse_press)
        # 连接滚轮缩放事件
        self.canvas1.mpl_connect('scroll_event', self.on_scroll)
        self.canvas2.mpl_connect('scroll_event', self.on_scroll)
        # 连接平移相关事件（鼠标移动与释放）
        self.canvas1.mpl_connect('motion_notify_event', self.on_mouse_move)
        self.canvas1.mpl_connect('button_release_event', self.on_mouse_release)
        self.canvas2.mpl_connect('motion_notify_event', self.on_mouse_move)
        self.canvas2.mpl_connect('button_release_event', self.on_mouse_release)
        
        # 初始化状态变量
        self.coord_tooltip = None
        # 平移(pan)状态变量
        self.is_panning = False
        self.pan_ax = None
        self.pan_start_x = None
        self.pan_start_y = None
        self.pan_orig_xlim = None
        self.pan_orig_ylim = None
        # 使用左键（mouse button 1）启动平移
        self.pan_button = 1

        # 在设置完UI后刷新画布
        self.figure1.tight_layout()
        self.figure2.tight_layout()
        # 创建曲线选择按钮（如果有多个传感器列）
        self.create_curve_buttons()
    
    def apply_points_settings(self):
        """应用左右检测点数设置"""
        try:
            left_points = int(self.left_points_edit.text())
            right_points = int(self.right_points_edit.text())
            
            if left_points >= 0 and right_points >= 0:
                self.left_points = left_points
                self.right_points = right_points
                
                # 重新显示当前段（如果有选中的话）
                # 修复：使用正确的属性名current_segment_index而不是_current_segment_index
                if self.current_segment_index >= 0:
                    self.show_segment(self.current_segment_index)
            else:
                from PyQt5.QtWidgets import QMessageBox
                QMessageBox.warning(self, "警告", "检测点数必须为非负整数！")
        except ValueError:
            from PyQt5.QtWidgets import QMessageBox
            QMessageBox.warning(self, "警告", "请输入有效的整数！")
    
    def is_click_on_curve(self, x, y, ax, x_data, y_data, pixel_threshold=10):
        """检查点击是否在曲线上"""
        if not x_data or not ax:
            return False, None
        
        try:
            # 转换为像素坐标
            points = np.column_stack([x_data, y_data])
            pixels = ax.transData.transform(points)
            
            # 点击位置转换为像素坐标
            click_pixel = ax.transData.transform((x, y))
            click_x_pixel, click_y_pixel = click_pixel
            
            # 计算距离
            distances = np.sqrt((pixels[:, 0] - click_x_pixel)**2 + (pixels[:, 1] - click_y_pixel)**2)
            min_dist = np.min(distances)
            nearest_idx = np.argmin(distances)
            
            return (True, nearest_idx) if min_dist <= pixel_threshold else (False, None)
        except Exception:
            return False, None
    
    def show_coord_tooltip(self, x, y, x_click=None, y_click=None):
        """显示坐标提示框"""
        self.hide_coord_tooltip()
        
        try:
            from PyQt5.QtWidgets import QLabel
            from PyQt5.QtGui import QCursor
            
            self.coord_tooltip = QLabel(f"x={x:.0f}, y={y:.4f}", self)
            self.coord_tooltip.setStyleSheet("""
                background-color: #333333;
                color: #FFFFFF;
                border: 1px solid #666666;
                border-radius: 4px;
                padding: 5px;
                font-weight: bold;
            """)
            self.coord_tooltip.setWindowFlags(Qt.ToolTip)
            
            # 定位提示框
            tooltip_width = self.coord_tooltip.sizeHint().width()
            tooltip_height = self.coord_tooltip.sizeHint().height()
            
            # 如果提供了点击坐标，则使用点击位置作为提示框的基准位置
            if x_click is not None and y_click is not None:
                # 获取光标位置
                cursor_pos = QCursor.pos()
                # 将光标位置转换为窗口坐标
                pos = self.mapFromGlobal(cursor_pos)
                x_pos = pos.x()
                y_pos = pos.y()
            else:
                # 如果没有提供点击坐标，使用光标位置
                pos = self.mapFromGlobal(QCursor.pos())
                x_pos = pos.x()
                y_pos = pos.y()
            
            # 确保提示框不会超出窗口边界
            x_pos = max(10, min(x_pos + 10, self.width() - tooltip_width - 10))
            y_pos = max(10, min(y_pos + 10, self.height() - tooltip_height - 10))
            
            self.coord_tooltip.move(x_pos, y_pos)
            self.coord_tooltip.show()
        except Exception:
            pass
    
    def hide_coord_tooltip(self):
        """隐藏坐标提示框"""
        if self.coord_tooltip:
            try:
                self.coord_tooltip.hide()
                self.coord_tooltip = None
            except:
                self.coord_tooltip = None

    def on_scroll(self, event):
        """处理滚轮缩放事件：以鼠标指向的点为中心对当前坐标轴进行水平方向缩放"""
        try:
            ax = event.inaxes
            if ax is None:
                return

            # zoom in when scrolling up, zoom out when scrolling down
            base_scale = 0.4
            # event.step may not be present in older mpl; use event.button 'up'/'down'
            try:
                direction = 1 if getattr(event, 'step', 0) > 0 else -1
            except Exception:
                direction = 1 if getattr(event, 'button', '') == 'up' else -1

            scale_factor = base_scale if direction > 0 else 1.0 / base_scale

            # Get mouse position in data coordinates
            xdata = event.xdata
            ydata = event.ydata

            # If no data coords, use axis center
            if xdata is None or ydata is None:
                xlim = ax.get_xlim()
                ylim = ax.get_ylim()
                xdata = (xlim[0] + xlim[1]) / 2.0
                ydata = (ylim[0] + ylim[1]) / 2.0

            # Compute new limits (仅在X轴方向进行缩放)
            x_left, x_right = ax.get_xlim()
            y_bottom, y_top = ax.get_ylim()

            new_width_left = (xdata - x_left) * scale_factor
            new_width_right = (x_right - xdata) * scale_factor
            new_x_left = xdata - new_width_left
            new_x_right = xdata + new_width_right

            # 保持Y轴范围不变，仅缩放X轴
            ax.set_xlim(new_x_left, new_x_right)
            # 不再修改Y轴范围
            # ax.set_ylim(new_y_bottom, new_y_top)

            # 同步同一splitter中的另一个图表
            sync_ax = None
            target_canvas = None
            if ax == self.ax1_cap:
                sync_ax = self.ax1_state
                target_canvas = self.canvas1
            elif ax == self.ax1_state:
                sync_ax = self.ax1_cap
                target_canvas = self.canvas1
            elif ax == self.ax2_cap:
                sync_ax = self.ax2_state
                target_canvas = self.canvas2
            elif ax == self.ax2_state:
                sync_ax = self.ax2_cap
                target_canvas = self.canvas2

            # 如果找到需要同步的轴，则同步x轴范围
            if sync_ax is not None:
                sync_ax.set_xlim(new_x_left, new_x_right)

            # redraw the appropriate canvas
            try:
                if ax in (self.ax1_cap, self.ax1_state):
                    self.canvas1.draw_idle()
                elif ax in (self.ax2_cap, self.ax2_state):
                    self.canvas2.draw_idle()
                else:
                    # fallback: redraw both
                    self.canvas1.draw_idle()
                    self.canvas2.draw_idle()
            except Exception:
                self.canvas1.draw()
                self.canvas2.draw()
        except Exception:
            # swallow exceptions to avoid breaking interaction
            pass
    
    def on_mouse_press(self, event):
        """处理鼠标按下事件"""
        self.hide_coord_tooltip()
        
        # 验证事件
        if not event.inaxes or event.xdata is None or event.ydata is None:
            return
        # 左键启动平移
        if event.button == self.pan_button:
            try:
                self.is_panning = True
                self.pan_ax = event.inaxes
                self.pan_start_x = event.xdata
                self.pan_start_y = event.ydata
                self.pan_orig_xlim = self.pan_ax.get_xlim()
                self.pan_orig_ylim = self.pan_ax.get_ylim()
            except Exception:
                self.is_panning = False
            return

        # 右键显示坐标
        if event.button == 3:  # 右键
            # 判断是哪个图表被点击
            if event.inaxes == self.ax1_cap:
                # 第一个图表的电容值曲线（支持多列）
                # 优先检查用户选中的曲线以便专注分析
                series_order = []
                # selected curves are stored as actual column indices
                for si, col_idx in enumerate(self.capacitor_columns):
                    if col_idx in self.selected_curves:
                        series_order.append((si, col_idx))
                for si, col_idx in enumerate(self.capacitor_columns):
                    if col_idx not in self.selected_curves:
                        series_order.append((si, col_idx))

                for si, col_idx in series_order:
                    series = self.capacitor_data_list[si]
                    on_curve, idx = self.is_click_on_curve(event.xdata, event.ydata,
                                                           self.ax1_cap,
                                                           list(range(len(series))),
                                                           series)
                    if on_curve and idx < len(series):
                        x_data = idx
                        y_data = series[idx]
                        label = self.headers[col_idx] if self.headers and col_idx < len(self.headers) else f"列{col_idx+1}"
                        self.show_coord_tooltip(x_data, y_data, event.xdata, event.ydata)
                        break
            elif event.inaxes == self.ax1_state:
                # 第一个图表的状态变量曲线
                on_curve, idx = self.is_click_on_curve(event.xdata, event.ydata, 
                                                       self.ax1_state, 
                                                       list(range(len(self.state_data))), 
                                                       self.state_data)
                if on_curve and idx < len(self.state_data):
                    x_data = idx
                    y_data = self.state_data[idx]
                    self.show_coord_tooltip(x_data, y_data, event.xdata, event.ydata)
            elif event.inaxes == self.ax2_cap:
                # 获取当前显示的段数据
                if hasattr(self, 'current_segment_x_range'):
                    # 第二个图表的电容值曲线（多列），优先选中曲线
                    series_order = []
                    for si, col_idx in enumerate(self.capacitor_columns):
                        if col_idx in self.selected_curves:
                            series_order.append((si, col_idx))
                    for si, col_idx in enumerate(self.capacitor_columns):
                        if col_idx not in self.selected_curves:
                            series_order.append((si, col_idx))

                    for si, col_idx in series_order:
                        series = self.capacitor_data_list[si]
                        y_segment = [series[i] for i in self.current_segment_x_range]
                        on_curve, idx = self.is_click_on_curve(event.xdata, event.ydata,
                                                               self.ax2_cap,
                                                               self.current_segment_x_range,  # 修复：使用正确的x轴数据
                                                               y_segment)
                        if on_curve and idx < len(self.current_segment_x_range):
                            x_data = self.current_segment_x_range[idx]
                            y_data = series[x_data]
                            label = self.headers[col_idx] if self.headers and col_idx < len(self.headers) else f"列{col_idx+1}"
                            self.show_coord_tooltip(x_data, y_data, event.xdata, event.ydata)
                            break
            elif event.inaxes == self.ax2_state:
                # 获取当前显示的段数据
                if hasattr(self, 'current_segment_x_range'):
                    # 第二个图表的状态变量曲线
                    on_curve, idx = self.is_click_on_curve(event.xdata, event.ydata, 
                                                           self.ax2_state, 
                                                           self.current_segment_x_range, 
                                                           [self.state_data[i] for i in self.current_segment_x_range])
                    if on_curve and idx < len(self.current_segment_x_range):
                        x_data = self.current_segment_x_range[idx]
                        y_data = self.state_data[x_data]
                        self.show_coord_tooltip(x_data, y_data, event.xdata, event.ydata)
            return

    def on_mouse_move(self, event):
        """处理鼠标移动事件：用于水平方向平移"""
        try:
            if not getattr(self, 'is_panning', False):
                return

            if self.pan_ax is None or event.inaxes is None:
                return

            # 需要有效的数据坐标
            if event.xdata is None or event.ydata is None:
                return

            # 计算偏移（以数据坐标为单位），只考虑X轴方向的移动
            dx = self.pan_start_x - event.xdata
            # dy = self.pan_start_y - event.ydata  # 不再使用Y轴方向的移动

            new_xlim = (self.pan_orig_xlim[0] + dx, self.pan_orig_xlim[1] + dx)
            # new_ylim = (self.pan_orig_ylim[0] + dy, self.pan_orig_ylim[1] + dy)  # 不再修改Y轴范围

            try:
                self.pan_ax.set_xlim(new_xlim)
                # 不再修改Y轴范围: self.pan_ax.set_ylim(new_ylim)
            except Exception:
                pass

            # 同步同一splitter中的另一个图表
            sync_ax = None
            if self.pan_ax == self.ax1_cap:
                sync_ax = self.ax1_state
            elif self.pan_ax == self.ax1_state:
                sync_ax = self.ax1_cap
            elif self.pan_ax == self.ax2_cap:
                sync_ax = self.ax2_state
            elif self.pan_ax == self.ax2_state:
                sync_ax = self.ax2_cap

            # 如果找到需要同步的轴，则同步x轴范围
            if sync_ax is not None:
                sync_ax.set_xlim(new_xlim)

            # 重绘对应的画布
            try:
                if self.pan_ax in (self.ax1_cap, self.ax1_state):
                    self.canvas1.draw_idle()
                else:
                    self.canvas2.draw_idle()
            except Exception:
                try:
                    self.canvas1.draw()
                    self.canvas2.draw()
                except Exception:
                    pass
        except Exception:
            pass

    def on_mouse_release(self, event):
        """处理鼠标释放事件：停止平移"""
        try:
            if getattr(self, 'is_panning', False):
                self.is_panning = False
                self.pan_ax = None
                self.pan_start_x = None
                self.pan_start_y = None
                self.pan_orig_xlim = None
                self.pan_orig_ylim = None
                
                # 确保两个canvas都被重绘以保持同步
                self.canvas1.draw_idle()
                self.canvas2.draw_idle()
        except Exception:
            pass
    
    def on_pick_state_segment(self, event):
        """处理状态段标签点击事件"""
        # 检查是否点击了文本对象
        if isinstance(event.artist, matplotlib.text.Text):
            # 获取文本对象的自定义属性（状态段索引）
            segment_index = event.artist.get_gid()
            if segment_index is not None:
                # 跳转到对应的状态段，但不重置第一个splitter的视图
                self.show_segment_without_reset(segment_index)
    
    def keyPressEvent(self, event):
        """处理键盘事件，支持方向键切换状态段"""
        # 确保事件不会被其他控件处理
        # 按 R 键复位视图
        if event.key() == Qt.Key_R:
            try:
                self.reset_view()
            except Exception:
                pass
            event.accept()
            return

        if self.state_segments and event.key() in (Qt.Key_Up, Qt.Key_Down, Qt.Key_Left, Qt.Key_Right):
            if event.key() == Qt.Key_Up or event.key() == Qt.Key_Left:
                self.navigate_segments(-1)
            elif event.key() == Qt.Key_Down or event.key() == Qt.Key_Right:
                self.navigate_segments(1)
            event.accept()  # 标记事件已被处理
            return
        super().keyPressEvent(event)
    
    def keyReleaseEvent(self, event):
        """处理键盘释放事件"""
        # 空实现，但确保事件不会向上传播
        event.accept()
    
    def navigate_segments(self, direction):
        """导航到下一个或上一个状态段"""
        if not self.state_segments:
            return
            
        if self.current_segment_index == -1:
            # 如果还没有选中任何段，默认选择第一个
            if direction > 0:  # 向下导航
                new_index = 0
            else:  # 向上导航
                new_index = len(self.state_segments) - 1
        else:
            # 计算新的索引，实现循环导航
            new_index = (self.current_segment_index + direction) % len(self.state_segments)
            
        self.show_segment(new_index)

    def reset_view(self):
        """将所有图表恢复到默认位置/缩放：重新绘制全图与当前段（触发自动缩放逻辑）。"""
        try:
            # 重新绘制全图（其中包含基于选中曲线或所有曲线的自适应缩放）
            self._plot_all_data()
        except Exception:
            pass

        # 如果有当前段，重新绘制段图（同样包含自适应缩放）
        try:
            if self.current_segment_index >= 0 and self.state_segments:
                seg = self.state_segments[self.current_segment_index]
                self._plot_segment_data(seg)
        except Exception:
            pass
    
    def show_segment(self, segment_index):
        """显示指定状态段的数据"""
        if not self.state_segments or segment_index >= len(self.state_segments):
            return
            
        # 保存当前段索引
        self.current_segment_index = segment_index
        segment = self.state_segments[segment_index]
        
        # 更新下拉按钮文本
        self.segment_dropdown.setText(f"状态段 {segment_index+1} (连1点数: {segment['length']})")
        
        # 更新菜单项的选中状态
        for i, action in enumerate(self.segment_actions):
            if i == segment_index:
                action.setFont(QFont(action.font().family(), action.font().pointSize(), QFont.Bold))
            else:
                action.setFont(QFont(action.font().family(), action.font().pointSize(), QFont.Normal))
        
        # 绘制第一个splitter - 所有数据
        self._plot_all_data()
        
        # 绘制第二个splitter - 当前段及其周边数据
        self._plot_segment_data(segment)
        
        # 计算并显示统计信息
        self._calculate_and_show_stats(segment)
        
    def show_segment_without_reset(self, segment_index):
        """显示指定状态段的数据，但不重置第一个splitter的视图"""
        if not self.state_segments or segment_index >= len(self.state_segments):
            return
            
        # 保存当前段索引
        self.current_segment_index = segment_index
        segment = self.state_segments[segment_index]
        
        # 更新下拉按钮文本
        self.segment_dropdown.setText(f"状态段 {segment_index+1} (连1点数: {segment['length']})")
        
        # 更新菜单项的选中状态
        for i, action in enumerate(self.segment_actions):
            if i == segment_index:
                action.setFont(QFont(action.font().family(), action.font().pointSize(), QFont.Bold))
            else:
                action.setFont(QFont(action.font().family(), action.font().pointSize(), QFont.Normal))
        
        # 保存当前第一个splitter的视图范围
        cap_xlim = self.ax1_cap.get_xlim()
        cap_ylim = self.ax1_cap.get_ylim()
        state_xlim = self.ax1_state.get_xlim()
        state_ylim = self.ax1_state.get_ylim()
        
        # 绘制第一个splitter - 所有数据（但保持原有视图范围）
        self._plot_all_data()
        
        # 恢复第一个splitter的视图范围
        self.ax1_cap.set_xlim(cap_xlim)
        self.ax1_cap.set_ylim(cap_ylim)
        self.ax1_state.set_xlim(state_xlim)
        self.ax1_state.set_ylim(state_ylim)
        
        # 绘制第二个splitter - 当前段及其周边数据
        self._plot_segment_data(segment)
        
        # 计算并显示统计信息
        self._calculate_and_show_stats(segment)
        
        # 重绘第一个splitter以反映变化
        self.canvas1.draw_idle()
    
    def _plot_all_data(self):
        """绘制所有数据曲线"""
        # 清除之前的绘图
        self.ax1_cap.clear()
        self.ax1_state.clear()
        
        # 绘制电容值曲线（多列）
        for si, series in enumerate(self.capacitor_data_list):
            col_idx = self.capacitor_columns[si]
            label = self.headers[col_idx] if self.headers and col_idx < len(self.headers) else f"列{col_idx+1}"
            # 如果用户选择了专注曲线，则突出显示选中曲线，其他曲线淡化
            if self.selected_curves:
                if col_idx in self.selected_curves:
                    self.ax1_cap.plot(series, linewidth=1.5, label=label)
                else:
                    self.ax1_cap.plot(series, linewidth=1, label=label, alpha=0.3)
            else:
                self.ax1_cap.plot(series, linewidth=1, label=label)
        self.ax1_cap.set_title("传感器值曲线", fontsize=10)
        self.ax1_cap.set_ylabel("传感器值", fontsize=9)
        self.ax1_cap.grid(True)
        self.ax1_cap.legend(fontsize=9)
        self.ax1_cap.tick_params(labelsize=8)
        # 禁用科学计数法
        self.ax1_cap.ticklabel_format(style='plain', axis='y')
        
        # 绘制状态变量曲线
        self.ax1_state.plot(self.state_data, 'r-', linewidth=1)  # 移除label参数
        self.ax1_state.set_title("状态变量曲线", fontsize=10)
        self.ax1_state.set_xlabel("数据点", fontsize=9)
        self.ax1_state.set_ylabel("状态值", fontsize=9)
        self.ax1_state.grid(True)
        # 移除legend显示
        self.ax1_state.tick_params(labelsize=8)
        # 禁用科学计数法
        self.ax1_state.ticklabel_format(style='plain', axis='y')
        
        # 标注状态段
        for i, segment in enumerate(self.state_segments):
            self.ax1_state.axvspan(segment["start"], segment["end"], 
                                   alpha=0.3, color='yellow')
            # 在每个状态段中间位置显示状态段编号
            mid_point = (segment["start"] + segment["end"]) / 2
            # 在状态值为1的位置显示状态段编号，添加picker属性以支持点击事件
            text = self.ax1_state.text(mid_point, 0.5, f'状态段{i+1}', 
                                ha='center', va='center', fontsize=9,
                                bbox=dict(boxstyle="round,pad=0.2", facecolor="white", alpha=0.7),
                                picker=True)
            # 为文本标签添加自定义属性，存储对应的状态段索引
            text.set_gid(i)
        
        # 连接点击事件处理器
        self.canvas1.mpl_connect('pick_event', self.on_pick_state_segment)
        
        # 根据当前可见曲线自适应坐标轴（如果有选中曲线则仅基于选中曲线，否则基于所有曲线）
        try:
            # 选择要用于缩放的曲线索引（表格列索引）
            if self.selected_curves:
                cols_to_use = [col for col in self.capacitor_columns if col in self.selected_curves]
            else:
                cols_to_use = list(self.capacitor_columns)

            all_x = []
            all_y = []
            for si, col_idx in enumerate(self.capacitor_columns):
                if col_idx in cols_to_use:
                    series = self.capacitor_data_list[si]
                    x = list(range(len(series)))
                    y = [v for v in series]
                    if len(x) > 0 and len(y) > 0:
                        all_x.extend(x)
                        # 过滤 nan 值
                        try:
                            all_y.extend([v for v in y if not np.isnan(v)])
                        except Exception:
                            all_y.extend(y)

            if all_x and all_y:
                x_min, x_max = min(all_x), max(all_x)
                y_min, y_max = min(all_y), max(all_y)
                x_margin = max(1, (x_max - x_min) * 0.05)
                y_margin = (y_max - y_min) * 0.05 if (y_max - y_min) != 0 else 0.5
                self.ax1_cap.set_xlim(x_min - x_margin, x_max + x_margin)
                self.ax1_cap.set_ylim(y_min - y_margin, y_max + y_margin)
        except Exception:
            pass

        self.canvas1.draw()
    
    def _plot_segment_data(self, segment):
        """绘制指定段及其周边数据"""
        # 清除之前的绘图
        self.ax2_cap.clear()
        self.ax2_state.clear()
        
        # 计算左右两侧连续0点的数量
        left_zeros = self._count_consecutive_zeros(segment["start"] - 1, -1, -1)
        right_zeros = self._count_consecutive_zeros(segment["end"] + 1, len(self.state_data), 1)
        
        # 确定实际要显示的点数（使用用户自定义的点数，但不超过实际连续0点数）
        display_points_left = min(self.left_points, left_zeros)
        display_points_right = min(self.right_points, right_zeros)

        # 计算显示范围（基于 state_data 的长度，所有序列已按 state 对齐）
        start_idx = max(0, segment["start"] - display_points_left)
        end_idx = min(len(self.state_data) - 1, segment["end"] + display_points_right)
        
        # 提取数据
        x_range = list(range(start_idx, end_idx + 1))
        cap_values_list = [series[start_idx:end_idx + 1] for series in self.capacitor_data_list]
        state_values = self.state_data[start_idx:end_idx + 1]
        
        # 保存当前显示的x_range，用于鼠标点击检测
        self.current_segment_x_range = x_range
        
        # 绘制电容值曲线（多列）
        for si, cap_values in enumerate(cap_values_list):
            col_idx = self.capacitor_columns[si]
            label = self.headers[col_idx] if self.headers and col_idx < len(self.headers) else f"列{col_idx+1}"
            # 突出显示选中曲线
            if self.selected_curves:
                if col_idx in self.selected_curves:
                    self.ax2_cap.plot(x_range, cap_values, linewidth=1.5, marker='o', markersize=3, label=label)
                else:
                    self.ax2_cap.plot(x_range, cap_values, linewidth=1, marker='o', markersize=2, label=label, alpha=0.3)
            else:
                self.ax2_cap.plot(x_range, cap_values, linewidth=1, marker='o', markersize=2, label=label)
        self.ax2_cap.set_title(f"状态段 {self.state_segments.index(segment)+1} 及周边数据 (点 {start_idx} 到 {end_idx})", fontsize=10)
        self.ax2_cap.set_ylabel("传感器值", fontsize=9)
        self.ax2_cap.grid(True)

        # 用不同颜色标示状态段
        self.ax2_cap.axvspan(segment["start"], segment["end"], 
                             alpha=0.3, color='yellow', label="选中状态段")
        self.ax2_cap.legend(fontsize=9)
        self.ax2_cap.tick_params(labelsize=8)
        # 禁用科学计数法
        self.ax2_cap.ticklabel_format(style='plain', axis='y')
        
        # 绘制状态变量曲线
        self.ax2_state.plot(x_range, state_values, 'r-', linewidth=1, marker='o', markersize=2)  # 移除label参数
        self.ax2_state.set_title("状态变量", fontsize=10)
        self.ax2_state.set_xlabel("数据点", fontsize=9)
        self.ax2_state.set_ylabel("状态值", fontsize=9)
        self.ax2_state.grid(True)
        self.ax2_state.set_ylim(min(state_values) - 0.5, max(state_values) + 0.5)
        
        # 用不同颜色标示状态段
        self.ax2_state.axvspan(segment["start"], segment["end"], 
                               alpha=0.3, color='yellow')  # 移除label参数
        # 不再显示legend
        self.ax2_state.tick_params(labelsize=8)
        # 禁用科学计数法
        self.ax2_state.ticklabel_format(style='plain', axis='y')
        
        # 自适应坐标轴：基于当前显示的（或选中）曲线调整范围
        try:
            # x_range 已是显示的索引
            if x_range:
                x_min, x_max = min(x_range), max(x_range)
                x_margin = max(1, (x_max - x_min) * 0.02)
                self.ax2_cap.set_xlim(x_min - x_margin, x_max + x_margin)

            # 选取用于缩放的曲线的数据
            if self.selected_curves:
                use_indices = [i for i, col in enumerate(self.capacitor_columns) if col in self.selected_curves]
            else:
                use_indices = list(range(len(cap_values_list)))

            all_y = []
            for ui in use_indices:
                seg = cap_values_list[ui]
                try:
                    all_y.extend([v for v in seg if not np.isnan(v)])
                except Exception:
                    all_y.extend(seg)

            if all_y:
                y_min, y_max = min(all_y), max(all_y)
                y_margin = (y_max - y_min) * 0.05 if (y_max - y_min) != 0 else 0.5
                self.ax2_cap.set_ylim(y_min - y_margin, y_max + y_margin)
        except Exception:
            pass

        self.canvas2.draw()
    
    def _count_consecutive_zeros(self, start_idx, end_idx, step):
        """
        计算从start_idx开始，以step方向，连续为0的点的数量
        Args:
            start_idx: 起始索引
            end_idx: 结束索引
            step: 方向步长 (-1表示向左，1表示向右)
        Returns:
            连续为0的点的数量
        """
        count = 0
        i = start_idx
        
        while (step == -1 and i >= end_idx) or (step == 1 and i < end_idx):
            if 0 <= i < len(self.state_data) and self.state_data[i] == 0:
                count += 1
                i += step
            else:
                break
                
        return count
    
    def _calculate_stats_for_export(self):
        """为导出计算所有状态段的统计信息"""
        all_segments_stats = []
        
        # 遍历所有状态段
        for segment in self.state_segments:
            # 计算左右两侧连续0点的数量（与绘图一致）
            left_zeros = self._count_consecutive_zeros(segment["start"] - 1, -1, -1)
            right_zeros = self._count_consecutive_zeros(segment["end"] + 1, len(self.state_data), 1)
            
            # 确定实际用于统计的点数（使用用户自定义的点数，但不超过实际连续0点数）
            stat_points_left = min(self.left_points, left_zeros)
            stat_points_right = min(self.right_points, right_zeros)
            
            # 计算左右段的索引范围（与绘图保持一致）
            left_start = max(0, segment["start"] - stat_points_left)
            left_end = segment["start"]
            right_start = segment["end"] + 1
            right_end = min(len(self.state_data), segment["end"] + 1 + stat_points_right)
            
            # 为每个传感器列计算统计量
            sensors_stats = []
            for si, series in enumerate(self.capacitor_data_list):
                name = self.headers[self.capacitor_columns[si]] if self.capacitor_columns and si < len(self.capacitor_columns) and self.headers else f"列{self.capacitor_columns[si]+1}"
                
                # 连1段统计
                seg_vals = [v for v in series[segment["start"]:segment["end"] + 1] if not np.isnan(v)]
                if seg_vals:
                    seg_min = min(seg_vals)
                    seg_max = max(seg_vals)
                    seg_mean = np.mean(seg_vals)
                    seg_pp = seg_max - seg_min
                else:
                    seg_min = seg_max = seg_mean = seg_pp = np.nan
                
                # 左侧0段统计
                left_vals = [v for v in series[left_start:left_end] if not np.isnan(v)] if left_end > left_start else []
                if left_vals:
                    left_min = min(left_vals)
                    left_max = max(left_vals)
                    left_mean = np.mean(left_vals)
                    left_pp = left_max - left_min
                else:
                    left_min = left_max = left_mean = left_pp = np.nan
                
                # 右侧0段统计
                right_vals = [v for v in series[right_start:right_end] if not np.isnan(v)] if right_end > right_start else []
                if right_vals:
                    right_min = min(right_vals)
                    right_max = max(right_vals)
                    right_mean = np.mean(right_vals)
                    right_pp = right_max - right_min
                else:
                    right_min = right_max = right_mean = right_pp = np.nan
                
                sensors_stats.append({
                    'name': name,
                    'seg_min': seg_min, 
                    'seg_max': seg_max, 
                    'seg_mean': seg_mean, 
                    'seg_pp': seg_pp,
                    'left_min': left_min,
                    'left_max': left_max,
                    'left_mean': left_mean,
                    'left_pp': left_pp,
                    'right_min': right_min,
                    'right_max': right_max,
                    'right_mean': right_mean,
                    'right_pp': right_pp
                })
            
            all_segments_stats.append({
                'segment': segment,
                'sensors_stats': sensors_stats
            })
        
        return all_segments_stats
    
    def _calculate_and_show_stats(self, segment):
        """计算并显示统计信息"""
        # 计算左右两侧连续0点的数量（与绘图一致）
        left_zeros = self._count_consecutive_zeros(segment["start"] - 1, -1, -1)
        right_zeros = self._count_consecutive_zeros(segment["end"] + 1, len(self.state_data), 1)

        # 确定实际要显示的点数（使用用户自定义的点数，但不超过实际连续0点数）
        stat_points_left = min(self.left_points, left_zeros)
        stat_points_right = min(self.right_points, right_zeros)

        # 计算左右段的索引范围（与绘图保持一致）
        left_start = max(0, segment["start"] - stat_points_left)
        left_end = segment["start"]
        right_start = segment["end"] + 1
        right_end = min(len(self.state_data), segment["end"] + 1 + stat_points_right)

        # 为每个传感器列计算统计量，并拼接为HTML
        sensors_stats = []
        for si, series in enumerate(self.capacitor_data_list):
            name = self.headers[self.capacitor_columns[si]] if self.capacitor_columns and si < len(self.capacitor_columns) and self.headers else f"列{self.capacitor_columns[si]+1}"

            # 连1段统计
            seg_vals = [v for v in series[segment["start"]:segment["end"] + 1] if not np.isnan(v)]
            if seg_vals:
                seg_min = min(seg_vals)
                seg_max = max(seg_vals)
                seg_mean = np.mean(seg_vals)
                seg_pp = seg_max - seg_min
            else:
                seg_min = seg_max = seg_mean = seg_pp = np.nan

            # 左侧0段统计
            left_vals = [v for v in series[left_start:left_end] if not np.isnan(v)] if left_end > left_start else []
            if left_vals:
                left_min = min(left_vals)
                left_max = max(left_vals)
                left_mean = np.mean(left_vals)
                left_pp = left_max - left_min
            else:
                left_min = left_max = left_mean = left_pp = np.nan

            # 右侧0段统计
            right_vals = [v for v in series[right_start:right_end] if not np.isnan(v)] if right_end > right_start else []
            if right_vals:
                right_min = min(right_vals)
                right_max = max(right_vals)
                right_mean = np.mean(right_vals)
                right_pp = right_max - right_min
            else:
                right_min = right_max = right_mean = right_pp = np.nan

            sensors_stats.append({
                'name': name,
                'seg_min': seg_min, 'seg_max': seg_max, 'seg_mean': seg_mean, 'seg_pp': seg_pp,
                'left_min': left_min, 'left_max': left_max, 'left_mean': left_mean, 'left_pp': left_pp,
                'right_min': right_min, 'right_max': right_max, 'right_mean': right_mean, 'right_pp': right_pp
            })

        # 拼接统计信息HTML
        stats_parts = [f"<b>状态段信息:</b><br>起始点: {segment['start']}<br>结束点: {segment['end']}<br>段长: {segment['length']}<br><br>"]
        for s in sensors_stats:
            # 格式化数值，处理NaN情况
            def format_val(val):
                if np.isnan(val):
                    return "N/A"
                else:
                    return f"{val:.4f}"
            
            stats_parts.append(f"<b>传感器: {s['name']}</b><br>")
            stats_parts.append(f"<b>连1段:</b><br>最小值: {format_val(s['seg_min'])}<br>最大值: {format_val(s['seg_max'])}<br>均值: {format_val(s['seg_mean'])}<br>峰峰值: {format_val(s['seg_pp'])}<br>")
            stats_parts.append(f"<b>左侧0段(左侧共{left_zeros}点, 实统计{stat_points_left}点):</b><br>最小值: {format_val(s['left_min'])}<br>最大值: {format_val(s['left_max'])}<br>均值: {format_val(s['left_mean'])}<br>峰峰值: {format_val(s['left_pp'])}<br>")
            stats_parts.append(f"<b>右侧0段(右侧共{right_zeros}点, 实统计{stat_points_right}点):</b><br>最小值: {format_val(s['right_min'])}<br>最大值: {format_val(s['right_max'])}<br>均值: {format_val(s['right_mean'])}<br>峰峰值: {format_val(s['right_pp'])}<br><br>")

        stats_text = ''.join(stats_parts)
        self.stats_label.setHtml(stats_text)

    def create_curve_buttons(self):
        """在右侧区域创建曲线选择按钮，允许用户专注某一条或多条曲线"""
        # 清理已有按钮布局
        try:
            for i in reversed(range(self.curve_buttons_layout.count())):
                w = self.curve_buttons_layout.itemAt(i).widget()
                if w:
                    w.setParent(None)
        except Exception:
            pass

        self.curve_buttons = []

        # 全部曲线按钮
        all_btn = QPushButton("全部曲线")
        all_btn.clicked.connect(lambda: self.select_curve(None))
        self.curve_buttons_layout.addWidget(all_btn)
        self.curve_buttons.append(all_btn)

        # 为每个传感器列创建按钮
        for si, col_idx in enumerate(self.capacitor_columns):
            label = self.headers[col_idx] if self.headers and col_idx < len(self.headers) else f"列{col_idx+1}"
            btn = QPushButton(label)
            btn.clicked.connect(lambda checked, idx=col_idx: self.select_curve(idx))
            self.curve_buttons_layout.addWidget(btn)
            self.curve_buttons.append(btn)

        # 移除addStretch()，改用根据按钮数量设置合适大小
        # self.curve_buttons_layout.addStretch()
        
        # 根据按钮数量设置曲线选择区域的最小高度
        button_count = len(self.curve_buttons)
        curves_group = self.curve_buttons_widget.parent()
        if curves_group:
            # 每个按钮大约需要30像素高度，加上边距和间距
            preferred_height = max(100, button_count * 35 + 20)
            curves_group.setMinimumHeight(preferred_height)
            
        # 初始状态为全部曲线
        self.select_curve(None)

    def select_curve(self, curve_idx):
        """切换曲线选中状态；curve_idx为None表示全部/取消全部"""
        if curve_idx is None:
            # 如果当前已经全部选中或未选中，则切换为选中全部
            if len(self.selected_curves) == 0 or len(self.selected_curves) == len(self.capacitor_columns):
                # 选中全部
                self.selected_curves = set(self.capacitor_columns)
            else:
                # 否则清空选择（显示所有）
                self.selected_curves.clear()
        else:
            # 切换单列选中状态
            if curve_idx in self.selected_curves:
                self.selected_curves.remove(curve_idx)
            else:
                self.selected_curves.add(curve_idx)

        # 更新按钮样式
        for i, btn in enumerate(self.curve_buttons):
            try:
                if i == 0:
                    if len(self.selected_curves) == 0 or len(self.selected_curves) == len(self.capacitor_columns):
                        btn.setStyleSheet("background-color: lightblue;")
                    else:
                        btn.setStyleSheet("")
                else:
                    col = self.capacitor_columns[i-1]
                    if col in self.selected_curves:
                        btn.setStyleSheet("background-color: lightblue;")
                    else:
                        btn.setStyleSheet("")
            except Exception:
                pass

        # 重新绘制当前显示段和全图
        if self.current_segment_index >= 0:
            self.show_segment(self.current_segment_index)

    def save_stats_to_excel(self):
        """将统计量保存到Excel文件"""
        if not XLSX_AVAILABLE:
            from PyQt5.QtWidgets import QMessageBox
            QMessageBox.warning(self, "功能不可用", "缺少openpyxl库，无法保存Excel文件。")
            return
        
        # 获取统计信息
        all_segments_stats = self._calculate_stats_for_export()
        
        if not all_segments_stats:
            from PyQt5.QtWidgets import QMessageBox
            QMessageBox.information(self, "无数据", "没有状态段数据可供导出。")
            return
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "状态统计量"
        
        # 获取传感器名称列表
        sensor_names = []
        if all_segments_stats:
            sensor_names = [stat['name'] for stat in all_segments_stats[0]['sensors_stats']]
        
        num_sensors = len(sensor_names)
        
        # 写入表头
        # 第一行：传感器名，每15列合并一个单元格（左侧0段4列 + 1列空隙 + 连1段4列 + 1列空隙 + 右侧0段4列）
        col_index = 1
        for sensor_name in sensor_names:
            start_col = col_index
            end_col = col_index + 13  # 15列（4+1+4+1+4）
            # 合并单元格并写入传感器名
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            cell = ws.cell(row=1, column=start_col, value=sensor_name)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
            col_index += 15  # 12列数据 + 3列间隙
        
        # 第二行：左侧0段，连1段，右侧0段（每个段落4列，段落间增加1列空隙）
        col_index = 1
        segment_labels = ['左侧0段', '连1段', '右侧0段']
        for _ in sensor_names:
            # 左侧0段
            start_col = col_index
            end_col = col_index + 3
            ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
            cell = ws.cell(row=2, column=start_col, value=segment_labels[0])
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
            col_index += 4
            
            # 添加空列
            col_index += 1
            
            # 连1段
            start_col = col_index
            end_col = col_index + 3
            ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
            cell = ws.cell(row=2, column=start_col, value=segment_labels[1])
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
            col_index += 4
            
            # 添加空列
            col_index += 1
            
            # 右侧0段
            start_col = col_index
            end_col = col_index + 3
            ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
            cell = ws.cell(row=2, column=start_col, value=segment_labels[2])
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
            col_index += 4
            
            # 传感器之间的间隙列
            col_index += 1
            
        # 第三行：最小值、最大值、均值、峰峰值（每个段落4列，段落间增加1列空隙）
        col_index = 1
        for _ in sensor_names:
            # 对于每个段落（左侧0段、连1段、右侧0段）
            for _ in range(3):
                headers = ['最小值', '最大值', '均值', '峰峰值']
                for i, header in enumerate(headers):
                    cell = ws.cell(row=3, column=col_index + i, value=header)
                    cell.alignment = Alignment(horizontal="center")
                    cell.font = Font(bold=True)
                col_index += 4
                # 每个段落后添加空列，除了最后一个段落
                if _ < 2:
                    col_index += 1
            col_index += 1  # 传感器之间的间隙列
            
        # 从第四行开始写入数据：每个状态段一行
        for row_index, segment_data in enumerate(all_segments_stats, start=4):
            col_index = 1
            for sensor_stat in segment_data['sensors_stats']:
                # 写入该传感器在左侧0段的4个统计值
                left_stats_values = [
                    sensor_stat.get('left_min', np.nan),
                    sensor_stat.get('left_max', np.nan),
                    sensor_stat.get('left_mean', np.nan),
                    sensor_stat.get('left_pp', np.nan)
                ]
                
                # 写入该传感器在连1段的4个统计值
                seg_stats_values = [
                    sensor_stat.get('seg_min', np.nan),
                    sensor_stat.get('seg_max', np.nan),
                    sensor_stat.get('seg_mean', np.nan),
                    sensor_stat.get('seg_pp', np.nan)
                ]
                
                # 写入该传感器在右侧0段的4个统计值
                right_stats_values = [
                    sensor_stat.get('right_min', np.nan),
                    sensor_stat.get('right_max', np.nan),
                    sensor_stat.get('right_mean', np.nan),
                    sensor_stat.get('right_pp', np.nan)
                ]
                
                # 写入所有值，注意在段落之间添加空列
                all_values = left_stats_values + [''] + seg_stats_values + [''] + right_stats_values
                for i, value in enumerate(all_values):
                    # 处理NaN值
                    if value == '':
                        # 空列不需要特殊处理
                        ws.cell(row=row_index, column=col_index + i, value='')
                    elif np.isnan(value):
                        cell = ws.cell(row=row_index, column=col_index + i, value="N/A")
                    else:
                        cell = ws.cell(row=row_index, column=col_index + i, value=value)
                        cell.number_format = '0.00'  # 设置数字格式

                col_index += 15  # 修改为16，因为现在总共有12列数据 + 3列间隙
        
        # 设置列宽
        for col_index in range(1, num_sensors * 15):  # 修改为16
            ws.column_dimensions[get_column_letter(col_index)].width = 12
        
        # 确定文件保存路径
        from SETTINGS.paths import get_log_directory, ensure_directory_exists
        import os
        
        # 确保Log目录存在
        log_dir = get_log_directory()
        ensure_directory_exists(log_dir)
        
        # 生成文件名
        if self.original_file_path:
            # 基于原始文件名生成Excel文件名
            original_filename = os.path.basename(self.original_file_path)
            name_without_ext, _ = os.path.splitext(original_filename)
            excel_filename = f"{name_without_ext}-状态变量检测.xlsx"
        else:
            # 默认文件名
            excel_filename = "状态变量检测.xlsx"
        
        # 完整文件路径
        file_path = os.path.join(log_dir, excel_filename)
        
        # 处理重名文件
        if os.path.exists(file_path):
            name_without_ext, ext = os.path.splitext(excel_filename)
            counter = 1
            while os.path.exists(os.path.join(log_dir, f"{name_without_ext}({counter}){ext}")):
                counter += 1
            excel_filename = f"{name_without_ext}({counter}){ext}"
            file_path = os.path.join(log_dir, excel_filename)
        
        try:
            wb.save(file_path)
            from PyQt5.QtWidgets import QHBoxLayout, QPushButton, QLabel, QVBoxLayout, QDialog, QStyle
            from PyQt5.QtCore import Qt
            
            # 创建自定义对话框
            msg_box = QDialog(self)
            msg_box.setWindowTitle("保存成功")
            msg_box.setModal(True)
            msg_box.resize(500, 200)
            
            # 创建主布局
            layout = QVBoxLayout(msg_box)
            layout.setSpacing(15)
            layout.setContentsMargins(20, 20, 20, 20)
            
            # 创建内容布局（图标+文本）
            content_layout = QHBoxLayout()
            content_layout.setSpacing(10)
            
            # 添加信息图标
            icon_label = QLabel()
            icon = msg_box.style().standardIcon(QStyle.SP_MessageBoxInformation)
            icon_label.setPixmap(icon.pixmap(32, 32))
            content_layout.addWidget(icon_label)
            
            # 添加文本信息
            from PyQt5.QtGui import QFont

            text_label = QLabel(f"统计量已保存到:\n{file_path}")
            text_label.setWordWrap(True)
            content_layout.addWidget(text_label)
            content_layout.addStretch()
            
            layout.addLayout(content_layout)
            
            # 创建按钮布局
            button_layout = QHBoxLayout()
            button_layout.setContentsMargins(10, 0, 10, 0)
            
            # 左侧按钮 - 打开目录
            open_dir_btn = QPushButton("打开Log目录")
            
            # 右侧按钮 - 确定
            ok_btn = QPushButton("确定")
            ok_btn.setDefault(True)
            
            # 添加按钮到布局
            button_layout.addWidget(open_dir_btn)
            button_layout.addStretch()
            button_layout.addWidget(ok_btn)
            
            layout.addLayout(button_layout)
            
            # 连接按钮信号
            open_dir_btn.clicked.connect(lambda: (msg_box.accept(), self._open_file_location(file_path)))
            ok_btn.clicked.connect(msg_box.accept)
            
            msg_box.exec_()
        except Exception as e:
            from PyQt5.QtWidgets import QMessageBox
            QMessageBox.critical(self, "保存失败", f"保存文件时出错:\n{str(e)}")

    def _open_file_location(self, file_path):
        """
        打开文件所在目录并选中文件
        """
        import platform
        import subprocess
        import os
        
        try:
            system = platform.system()
            if system == "Windows":
                # Windows系统使用explorer命令选中文件
                subprocess.Popen(f'explorer /select,"{os.path.normpath(file_path)}"', shell=True)
            elif system == "Darwin":  # macOS
                # macOS使用open命令
                subprocess.Popen(["open", "-R", file_path])
            else:  # Linux
                # Linux打开目录即可
                directory = os.path.dirname(file_path)
                subprocess.Popen(["xdg-open", directory])
        except Exception as e:
            from PyQt5.QtWidgets import QMessageBox
            QMessageBox.warning(self, "打开目录失败", f"无法打开文件所在目录:\n{str(e)}")

